"""
Wi-Fi Heat Mapper - GUI Application
Replaces all CLI commands with a visual interface.
"""
import FreeSimpleGUI as sg
import os
import sys
import logging
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import numpy as np
from PIL import Image

# Ensure package is importable
if not getattr(sys, 'frozen', False):
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ── Stream Redirection for Windowed Mode ───────────────────────────
class NullWriter:
    def write(self, x): pass
    def flush(self): pass

if sys.stdout is None:
    sys.stdout = NullWriter()
if sys.stderr is None:
    sys.stderr = NullWriter()

# ── Logging Setup ──────────────────────────────────────────────────
LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(
    sys.executable if getattr(sys, 'frozen', False) else __file__
)), "whm_debug.log")

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
    ]
)
# Also log to stdout if it's available (cli mode)
if not isinstance(sys.stdout, NullWriter):
    logging.getLogger().addHandler(logging.StreamHandler(sys.stdout))

logger = logging.getLogger("whm_app")
logger.info("=== Wi-Fi Heat Mapper iniciado ===")
logger.info("Log salvo em: %s", LOG_FILE)

from wifi_heat_mapper.config import ConfigurationOptions
from wifi_heat_mapper.misc import (
    check_application, process_iw, save_json, load_json,
    get_ip_address_from_interface,
    get_property_from
)
from wifi_heat_mapper.gui import start_gui
from wifi_heat_mapper.graph import generate_graph
from wifi_heat_mapper.windows_wlan import get_wifi_metrics_windows, scan_all_networks
from wifi_heat_mapper import __version__

sg.theme('DarkBlue3')
TITLE = "Wi-Fi Heat Mapper"


# ── Utility Functions ──────────────────────────────────────────────




def detect_wifi():
    try:
        # Tenta descobrir qual interface está conectada
        from wifi_heat_mapper.windows_wlan import get_wlanapi, open_handle, get_interfaces, close_handle
        api = get_wlanapi()
        if api:
            h = open_handle(api)
            if h:
                ifaces = get_interfaces(api, h)
                close_handle(api, h)
                for iface in ifaces:
                    if iface['state'] == 1: # Connected
                        metrics = get_wifi_metrics_windows(iface['description'])
                        if metrics:
                            metrics['interface_name'] = iface['description']
                            return metrics
        
        # Fallback para nome padrão
        metrics = get_wifi_metrics_windows("Wi-Fi")
        if metrics:
            metrics['interface_name'] = "Wi-Fi"
        return metrics
    except Exception as e:
        logger.exception("Erro ao detectar Wi-Fi: %s", e)
        return None


def get_available_graphs(supported_modes):
    modes_with_base = supported_modes + ["base"]
    result = []
    for key, opts in ConfigurationOptions.configuration.items():
        if set(opts["mode"]).intersection(set(modes_with_base)):
            result.append((key, opts["description"]))
    return result


# ── Main Menu ──────────────────────────────────────────────────────

def main_menu():
    layout = [
        [sg.Text(TITLE, font=("Helvetica", 20), justification="center", expand_x=True)],
        [sg.Text(f"v{__version__} — Windows Edition", font=("Helvetica", 10),
                 justification="center", expand_x=True)],
        [sg.HorizontalSeparator()],
        [sg.Text("")],
        [sg.Button("Novo Projeto", size=(30, 2), key="-NEW-")],
        [sg.Text("")],
        [sg.Button("Continuar Projeto", size=(30, 2), key="-OPEN-")],
        [sg.Text("")],
        [sg.Button("Gerar Mapas de Calor", size=(30, 2), key="-PLOT-")],
        [sg.Text("")],
        [sg.Button("Gerar Relatório Acadêmico", size=(30, 2), key="-REPORT-", button_color=("white", "#2c3e50"))],
        [sg.Text("")],
        [sg.Button("Sair", size=(30, 1), key="-EXIT-")],
    ]
    window = sg.Window(TITLE, layout, element_justification="center", finalize=True)
    while True:
        event, _ = window.read()
        if event in (sg.WIN_CLOSED, "-EXIT-"):
            window.close()
            return None
        for key in ("-NEW-", "-OPEN-", "-PLOT-", "-REPORT-"):
            if event == key:
                window.close()
                return key


# ── Configuration Wizard (replaces CLI bootstrap) ─────────────────

def config_wizard():
    supported_modes = ["base"]
    wifi_info = detect_wifi()

    available_graphs = get_available_graphs(supported_modes)

    if wifi_info:
        wifi_text = (f"SSID: {wifi_info['ssid']}  |  "
                     f"Sinal: {wifi_info['signal_strength']} dBm  |  "
                     f"Canal: {wifi_info['channel']}")
        det_ssid = wifi_info['ssid']
        iface_name = wifi_info['interface_name']
    else:
        wifi_text = "Wi-Fi não detectado automaticamente. Verifique se está conectado."
        det_ssid = ""
        iface_name = "Wi-Fi"

    graph_rows = [[sg.Checkbox(desc, key=f"-G-{key}", default=True)]
                  for key, desc in available_graphs]

    layout = [
        [sg.Text("Configuração do Projeto", font=("Helvetica", 16))],
        [sg.HorizontalSeparator()],
        [sg.Frame("Rede Wi-Fi Conectada", [
            [sg.Text(wifi_text)],
            [sg.Text("Interface:"), sg.Input(iface_name, key="-IFACE-", size=(25, 1)),
             sg.Text("  SSID detectado:"), sg.Text(det_ssid, key="-SSID-TXT-")],
            [sg.Text("Nota: Todas as redes visíveis serão coletadas automaticamente em cada ponto.",
                     font=("Helvetica", 9), text_color="yellow")],
        ], expand_x=True)],

        [sg.Frame("Gráficos Desejados", [
            [sg.Column(graph_rows, scrollable=True, vertical_scroll_only=True, 
                       size=(550, 250), key="-GRAPH-COL-")]
        ], expand_x=True)],
        [sg.Frame("Planta do Local", [
            [sg.Input(key="-MAP-", size=(45, 1)),
             sg.FileBrowse("Procurar", file_types=(("Imagens", "*.jpg *.jpeg *.png *.bmp"),))],
        ], expand_x=True)],
        [sg.Frame("Salvar Configuração Como", [
            [sg.Input("config.json", key="-SAVE-", size=(45, 1)),
             sg.FileSaveAs("Procurar", file_types=(("JSON", "*.json"),), default_extension=".json")],
        ], expand_x=True)],
        [sg.Button("Salvar e Iniciar Benchmark", key="-GO-", size=(28, 1)),
         sg.Button("Apenas Salvar", key="-SAVEONLY-", size=(15, 1)),
         sg.Button("Cancelar", key="-CANCEL-")],
    ]

    window = sg.Window(f"{TITLE} - Novo Projeto", layout, finalize=True)
    result = None

    while True:
        event, values = window.read()
        if event in (sg.WIN_CLOSED, "-CANCEL-"):
            break

        if event in ("-GO-", "-SAVEONLY-"):
            iface = values["-IFACE-"].strip()
            floor_map = values["-MAP-"].strip()
            save_path = values["-SAVE-"].strip()

            if not iface:
                sg.popup_error("Informe a interface de rede."); continue
            if not floor_map or not os.path.isfile(floor_map):
                sg.popup_error("Selecione uma planta válida."); continue
            if not save_path:
                sg.popup_error("Informe onde salvar a configuração."); continue

            sel_graphs = []
            sel_modes = set()
            for key, _ in available_graphs:
                if values.get(f"-G-{key}", False):
                    sel_graphs.append(key)
                    for r in ConfigurationOptions.configuration[key]["requirements"]:
                        sel_modes.add(r)
            if not sel_graphs:
                sg.popup_error("Selecione pelo menos um gráfico."); continue

            try:
                wifi_data = process_iw(iface)
                ssid = wifi_data["ssid"]
                bind_ip = get_ip_address_from_interface(iface)
            except Exception as e:
                sg.popup_error(f"Erro ao acessar a interface:\n{e}"); continue

            if not save_path.endswith(".json"):
                save_path += ".json"
            save_path = os.path.abspath(save_path)

            config_data = {
                "configuration": {
                    "graphs": sel_graphs,
                    "modes": list(sel_modes),
                    "backends": supported_modes + ["base"],
                    "version": __version__,
                    "target_interface": iface,
                    "adapter_description": iface,
                    "target_ip": bind_ip,
                    "ssid": ssid,
                    "speedtest": -1,
                    "libre-speed-list": "",
                    "benchmark_iterations": 1,
                },
                "results": {}
            }

            if not save_json(save_path, config_data):
                sg.popup_error("Erro ao salvar."); continue

            sg.popup_ok(f"Configuração salva em:\n{save_path}")

            if event == "-GO-":
                result = {"config": save_path, "map": os.path.abspath(floor_map)}
            break

    window.close()
    return result


# ── Open Existing Project ─────────────────────────────────────────

def open_project():
    layout = [
        [sg.Text("Continuar Projeto", font=("Helvetica", 16))],
        [sg.HorizontalSeparator()],
        [sg.Frame("Arquivos", [
            [sg.Text("Configuração (.json):"),
             sg.Input(key="-CFG-", size=(40, 1)),
             sg.FileBrowse("Procurar", file_types=(("JSON", "*.json"),))],
            [sg.Text("Planta do local:      "),
             sg.Input(key="-MAP-", size=(40, 1)),
             sg.FileBrowse("Procurar", file_types=(("Imagens", "*.jpg *.jpeg *.png *.bmp"),))],
        ], expand_x=True)],

        [sg.Button("Iniciar", key="-START-", size=(15, 1)),
         sg.Button("Cancelar", key="-CANCEL-")],
    ]
    window = sg.Window(f"{TITLE} - Continuar Projeto", layout, finalize=True)
    result = None
    while True:
        event, values = window.read()
        if event in (sg.WIN_CLOSED, "-CANCEL-"):
            break
        if event == "-START-":
            cfg = values["-CFG-"].strip()
            flr = values["-MAP-"].strip()
            if not cfg or not os.path.isfile(cfg):
                sg.popup_error("Selecione um .json válido."); continue
            if not flr or not os.path.isfile(flr):
                sg.popup_error("Selecione uma planta válida."); continue
            result = {"config": os.path.abspath(cfg), "map": os.path.abspath(flr)}
            break
    window.close()
    return result


# ── Interactive Heatmap Viewer ─────────────────────────────────────

def _extract_all_ssids(benchmark_results):
    """Extract all unique SSIDs from benchmark results."""
    ssids = set()
    for point in benchmark_results.values():
        nets = (point.get("results") or {}).get("networks", {})
        for ssid_name in nets:
            ssids.add(ssid_name)
    return sorted(ssids)


def _build_heatmap_figure(benchmark_results, floor_map_path, metric_key, selected_ssids, use_network_data, **kwargs):
    """Build a matplotlib Figure for the given metric and selected SSIDs."""
    import copy
    from scipy.interpolate import Rbf
    from matplotlib.pyplot import imread

    im = Image.open(floor_map_path)
    fdimx, fdimy = im.size

    config_entry = ConfigurationOptions.configuration.get(metric_key, {})
    vmin = config_entry.get("vmin", None)
    vmax = config_entry.get("vmax", None)
    reverse = config_entry.get("reverse", False)
    desc = config_entry.get("description", metric_key)

    xs, ys, zs = [], [], []
    raw_data = []
    for point_data in benchmark_results.values():
        results = point_data.get("results")
        if not results:
            continue
        px = point_data["position"]["x"]
        py = point_data["position"]["y"]

        # Se for métrica de rede mas nada estiver selecionado, não mostramos o "Target" (raiz)
        # Forçamos o usuário a selecionar o que ele quer ver na lista.
        if use_network_data and not selected_ssids:
            continue

        if use_network_data and selected_ssids:
            # Identifica a melhor rede (maior signal_strength) selecionada neste ponto de medição.
            # Isso garante que as métricas (sinal, SNR, canal, etc.) reflitam a rede à qual o cliente
            # se conectaria na prática, evitando distorções ou médias sem sentido físico (como médias de canais).
            nets = results.get("networks", {})
            best_ssid = None
            best_signal = -9999
            for ssid_name in selected_ssids:
                if ssid_name in nets:
                    sig = nets[ssid_name].get("signal_strength", -100)
                    if sig > best_signal:
                        best_signal = sig
                        best_ssid = ssid_name

            if best_ssid is None:
                # Nenhuma das redes selecionadas é visível neste ponto.
                # Definimos os piores valores físicos correspondentes à ausência de sinal
                # para que a interpolação RBF represente corretamente as áreas de sombra.
                if metric_key == "signal_strength":
                    val = -100
                elif metric_key == "snr_estimated":
                    val = 0
                elif metric_key == "signal_quality":
                    val = 0
                elif metric_key == "signal_quality_percent":
                    val = 0
                elif metric_key in ("channel", "channel_frequency"):
                    val = 0
                else:
                    val = 0
            elif metric_key not in nets[best_ssid]:
                continue
            else:
                val = nets[best_ssid][metric_key]
        else:
            if metric_key not in results:
                continue
            val = results[metric_key]

        xs.append(px)
        ys.append(py)
        zs.append(val)
        raw_data.append(point_data)

    if len(xs) < 3:
        # Not enough points for interpolation
        fig, ax = plt.subplots(1, 1, figsize=(fdimx / 100, fdimy / 100))
        ax.imshow(imread(floor_map_path)[::-1], interpolation='bilinear', zorder=1, alpha=1, origin="lower")
        ax.set_title("Dados insuficientes (mínimo 3 pontos)", fontsize=12)
        ax.axis('off')
        fig.tight_layout()
        return fig, xs, ys, zs, raw_data

    # Add boundary points
    xs += [0, 0, fdimx, fdimx]
    ys += [0, fdimy, fdimy, 0]
    # Busca limites fixos na configuração para cores consistentes (Estilo Profissional)
    config_entry = ConfigurationOptions.configuration.get(metric_key, {})
    if vmin is None:
        vmin = config_entry.get("vmin", min(zs))
    if vmax is None:
        vmax = config_entry.get("vmax", max(zs))
        
    reverse = config_entry.get("reverse", False)
    boundary_val = vmax if reverse else vmin
    zs += [boundary_val] * 4

    import math
    fdim_coef = math.sqrt(fdimx * fdimy)
    
    xi = np.linspace(0, fdimx, 100)
    yi = np.linspace(0, fdimy, 100)
    xi, yi = np.meshgrid(xi, yi)
    # Interpolação linear (RBF) — consistente com o módulo de exportação (graph.py)
    di = Rbf(xs, ys, zs, function="linear")
    zi = di(xi, yi)
    zi[zi < vmin] = vmin
    zi[zi > vmax] = vmax

    is_discrete = metric_key in ("co_channel_interference", "adjacent_channel_interference", "ap_density")
    if is_discrete and (math.ceil(vmax) - math.floor(vmin)) >= 1:
        levels_contour = np.arange(math.floor(vmin), math.ceil(vmax) + 1)
        levels_contourf = levels_contour
        fmt = '%d'
    else:
        levels_contourf = np.linspace(vmin, vmax, 31)
        levels_contour = np.linspace(vmin, vmax, 16)
        fmt = '%.1f'

    title_size = max(10, fdim_coef // 70)

    fig, ax = plt.subplots(1, 1, figsize=(fdimx / 100, fdimy / 100))
    # Heatmap preenchido com escala intuitiva Semáforo (Cores fortes e vibrantes)
    # Otimizado: levels reduzido de 100 para 30 e antialiased=False para suportar redimensionamento fluido
    cmap_name = "RdYlGn_r" if reverse else "RdYlGn"
    bench_plot = ax.contourf(xi, yi, zi, cmap=cmap_name, vmin=vmin, vmax=vmax,
                             alpha=0.60, zorder=150, antialiased=False, levels=levels_contourf, extend='both')
    
    # Adiciona Isolinhas (Contour Lines) para facilitar a leitura técnica
    contours = ax.contour(xi, yi, zi, levels=levels_contour, colors='black', 
                          linewidths=0.5, alpha=0.5, zorder=155)
    ax.clabel(contours, inline=True, fontsize=max(6, title_size // 2), fmt=fmt, colors='black')
    ms = max(6, fdim_coef // 150) # Aumentado para melhor visibilidade

    # AP Triangulation (Feature B)
    show_aps = kwargs.get("show_aps", True)
    if use_network_data and selected_ssids and show_aps:
        ap_locations = {}
        # Para cada rede selecionada, encontramos o ponto de maior sinal (Pico)
        for ssid in selected_ssids:
            max_val = -1000
            best_pos = None
            
            # Procuramos nos dados brutos o local onde esta rede específica foi mais forte
            for point_data in benchmark_results.values():
                nets = point_data.get("results", {}).get("networks", {})
                if ssid in nets:
                    val = nets[ssid].get(metric_key, -1000)
                    if val > max_val:
                        max_val = val
                        best_pos = point_data["position"]
            
            if best_pos:
                ap_locations[ssid] = best_pos

        # Plot estimated AP locations
        for ssid, pos in ap_locations.items():
            ax.plot(pos["x"], pos["y"], 'rx', markersize=max(10, ms*2), markeredgewidth=2, zorder=300)
            ax.text(pos["x"], pos["y"] + 5, f"AP: {ssid}", color='red', 
                    fontsize=max(8, title_size // 1.5), fontweight='bold', zorder=301,
                    bbox=dict(facecolor='white', alpha=0.8, edgecolor='none', pad=2))

    ax.imshow(imread(floor_map_path)[::-1], interpolation='bilinear', zorder=1, alpha=1, origin="lower")

    ax.plot(xs[:-4], ys[:-4], zorder=200, marker='o', markeredgecolor='black',
            markeredgewidth=0.5, linestyle='None', markersize=ms, label="Ponto de Medição")

    title_size = max(10, fdim_coef // 70)
    label_size = max(7, title_size - 5)
    cb = fig.colorbar(bench_plot)
    cb.ax.tick_params(labelsize=label_size)

    title = desc
    if use_network_data and selected_ssids:
        if len(selected_ssids) == 1:
            title += f" [{selected_ssids[0]}]"
        else:
            title += f" [{len(selected_ssids)} redes]"
    ax.set_title(title, fontsize=title_size)
    ax.axis('off')
    fig.tight_layout()
    return fig, xs, ys, zs, raw_data


def plot_dialog():
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

    def draw_figure(canvas, figure, xs, ys, zs, raw_data, selected_ssids):
        canvas_agg = FigureCanvasTkAgg(figure, canvas)
        canvas_agg.draw()
        canvas_agg.get_tk_widget().pack(side='top', fill='both', expand=1)
        
        # Tooltip interativo
        annot = figure.axes[0].annotate("", xy=(0,0), xytext=(20,20),
                                      textcoords="offset points",
                                      bbox=dict(boxstyle="round", fc="w", alpha=0.95),
                                      arrowprops=dict(arrowstyle="->"))
        annot.set_visible(False)

        def update_annot(ind, x, y, val, point_info):
            pos = [x, y]
            annot.xy = pos
            
            text = f"Valor: {val:.1f}\n"
            if point_info:
                ts = point_info.get("results", {}).get("timestamp", "")
                if ts: 
                    # Ex: 2026-05-11T16:50:00 -> 16:50:00
                    text += f"Hora: {ts[11:19]}\n"
                
                if selected_ssids and len(selected_ssids) > 0:
                    for ssid in selected_ssids[:3]: # Limit to 3 to avoid huge tooltip
                        net_info = point_info.get("results", {}).get("networks", {}).get(ssid, {})
                        if not net_info: continue
                        
                        vendor = net_info.get("vendor", "")
                        if vendor and vendor != "Unknown": 
                            text += f"[{ssid[:15]}] Vendor: {vendor}\n"
                            
                        sta = net_info.get("bss_load_station_count", 0)
                        util = net_info.get("bss_load_channel_utilization", 0.0)
                        if sta > 0 or util > 0:
                            text += f"[{ssid[:15]}] Load: {sta} cli / {util}% uso\n"
                            
            annot.set_text(text.strip())
            annot.get_bbox_patch().set_alpha(0.95)

        def hover(event):
            vis = annot.get_visible()
            if event.inaxes == figure.axes[0]:
                # Encontrar ponto mais próximo (nos dados originais xs, ys)
                cont_xs = xs[:-4] # Remove boundary points
                cont_ys = ys[:-4]
                cont_zs = zs[:-4]
                
                dists = np.sqrt((np.array(cont_xs)-event.xdata)**2 + (np.array(cont_ys)-event.ydata)**2)
                if len(dists) > 0:
                    min_idx = np.argmin(dists)
                    if dists[min_idx] < (fdim_coef / 30): # Raio de detecção
                        point_info = raw_data[min_idx] if min_idx < len(raw_data) else None
                        update_annot(min_idx, cont_xs[min_idx], cont_ys[min_idx], cont_zs[min_idx], point_info)
                        annot.set_visible(True)
                        figure.canvas.draw_idle()
                        return
            if vis:
                annot.set_visible(False)
                figure.canvas.draw_idle()

        figure.canvas.mpl_connect("motion_notify_event", hover)
        return canvas_agg

    def delete_figure(canvas_agg):
        canvas_agg.get_tk_widget().destroy()

    # Step 1: File selection
    file_layout = [
        [sg.Text("Visualizador Interativo de Heatmaps", font=("Helvetica", 16))],
        [sg.HorizontalSeparator()],
        [sg.Frame("Arquivos", [
            [sg.Text("Configuração (.json):"),
             sg.Input(key="-CFG-", size=(40, 1)),
             sg.FileBrowse("Procurar", file_types=(("JSON", "*.json"),))],
            [sg.Text("Planta do local:      "),
             sg.Input(key="-MAP-", size=(40, 1)),
             sg.FileBrowse("Procurar", file_types=(("Imagens", "*.jpg *.jpeg *.png *.bmp"),))],
        ], expand_x=True)],
        [sg.Button("Carregar", key="-LOAD-", size=(15, 1)),
         sg.Button("Cancelar", key="-CANCEL-")],
    ]
    file_win = sg.Window(f"{TITLE} - Selecionar Projeto", file_layout, finalize=True)
    cfg_path = flr_path = None
    while True:
        ev, vals = file_win.read()
        if ev in (sg.WIN_CLOSED, "-CANCEL-"):
            file_win.close()
            return
        if ev == "-LOAD-":
            cfg_path = vals["-CFG-"].strip()
            flr_path = vals["-MAP-"].strip()
            if not cfg_path or not os.path.isfile(cfg_path):
                sg.popup_error("Selecione um .json válido."); continue
            if not flr_path or not os.path.isfile(flr_path):
                sg.popup_error("Selecione uma planta válida."); continue
            break
    file_win.close()

    # Load data
    data = load_json(cfg_path)
    if not data:
        sg.popup_error("Erro ao carregar o arquivo de configuração.")
        return
    benchmark_results = get_property_from(data, "results")
    if not benchmark_results:
        sg.popup_error("Nenhum resultado de benchmark encontrado no arquivo.")
        return

    # Extract available SSIDs and metrics
    all_ssids = _extract_all_ssids(benchmark_results)
    
    secure_ssids = set()
    open_ssids = set()
    for point in benchmark_results.values():
        nets = point.get("results", {}).get("networks", {})
        for ssid, net_data in nets.items():
            if net_data.get("is_secure", False):
                secure_ssids.add(ssid)
            else:
                open_ssids.add(ssid)
                
    secure_ssids = sorted(list(secure_ssids))
    open_ssids = sorted(list(open_ssids))

    per_net_metrics = ["signal_strength", "snr_estimated", "signal_quality",
                       "signal_quality_percent", "channel"]
    metric_options = []
    for key, entry in ConfigurationOptions.configuration.items():
        metric_options.append((key, entry["description"]))

    # Build viewer layout with hide button
    def make_checkbox_row(ssid):
        return [sg.pin(sg.Column([[
            sg.Checkbox(ssid, key=f"-SSID-{ssid}", default=False, size=(45, 1)),
            sg.Button("✖", key=f"-HIDE-{ssid}", size=(2, 1), button_color=("white", "#c0392b"),
                      tooltip="Ocultar rede", border_width=0)
        ]], key=f"-ROW-{ssid}", pad=(0, 0)))]

    open_checkboxes = [make_checkbox_row(s) for s in open_ssids]
    secure_checkboxes = [make_checkbox_row(s) for s in secure_ssids]
    
    if not open_checkboxes:
        open_checkboxes = [[sg.Text("Nenhuma rede aberta encontrada", font=("Helvetica", 9, "italic"))]]
    if not secure_checkboxes:
        secure_checkboxes = [[sg.Text("Nenhuma rede fechada encontrada", font=("Helvetica", 9, "italic"))]]

    metric_names = [desc for _, desc in metric_options]
    metric_keys = [key for key, _ in metric_options]

    viewer_layout = [
        [sg.Column([
            [sg.Text("Métrica:", font=("Helvetica", 11, "bold")),
             sg.Text("ⓘ", font=("Helvetica", 14), key="-METRIC-HELP-", text_color="lightblue",
                     tooltip="Passe o mouse para ver a descrição da métrica")],
            [sg.Combo(metric_names, default_value=metric_names[0] if metric_names else "",
                      key="-METRIC-", size=(45, 1), enable_events=True, readonly=True)],
            [sg.HorizontalSeparator()],
            [sg.Text("Redes Abertas (Sem Senha):", font=("Helvetica", 10, "bold"), text_color="orange")],
            [sg.Column(open_checkboxes, scrollable=True, vertical_scroll_only=True,
                       size=(380, 160), key="-SSID-COL-OPEN-")],
            [sg.Text("Redes Fechadas (Seguras):", font=("Helvetica", 10, "bold"), text_color="lightgreen")],
            [sg.Column(secure_checkboxes, scrollable=True, vertical_scroll_only=True,
                       size=(380, 200), key="-SSID-COL-SECURE-")],
            [sg.Button("Selecionar Todas", key="-ALL-", size=(14, 1)),
             sg.Button("Limpar", key="-NONE-", size=(8, 1))],
            [sg.Button("Mostrar Redes Ocultas", key="-UNHIDE-", size=(25, 1), font=("Helvetica", 9),
                       button_color=("white", "#2980b9"))],
            [sg.Checkbox("Mostrar Localização de APs", default=True, key="-SHOW-APS-")],
            [sg.HorizontalSeparator()],
            [sg.Button("Atualizar Heatmap", key="-UPDATE-", size=(25, 2), font=("Helvetica", 11, "bold"), button_color=("white", "#27ae60"))],
            [sg.HorizontalSeparator()],
            [sg.Text("Exportar:", font=("Helvetica", 11, "bold"))],
            [sg.Combo(["png", "pdf", "svg"], default_value="png", key="-FMT-", size=(6, 1)),
             sg.Button("Salvar Imagem", key="-EXPORT-", size=(14, 1))],
            [sg.HorizontalSeparator()],
            [sg.Button("📏 Ver Escala Calibrada", key="-VIEW-SCALE-", size=(25, 1), font=("Helvetica", 10, "bold"), button_color=("white", "#7f8c8d"))],
        ], vertical_alignment='top', expand_y=True, size=(420, None)),
         sg.Column([
            [sg.Canvas(key="-CANVAS-", size=(800, 650))],
         ], expand_x=True, expand_y=True)],
    ]

    viewer_win = sg.Window(f"{TITLE} - Visualizador de Heatmaps", viewer_layout,
                           finalize=True, resizable=True, size=(1150, 750))

    # Initial variables
    current_metric_idx = 0
    current_metric = metric_keys[0] if metric_keys else "signal_strength"

    # Set initial help tooltip
    initial_help = ConfigurationOptions.configuration.get(current_metric, {}).get("help", "Sem descrição.")
    viewer_win["-METRIC-HELP-"].set_tooltip(initial_help)

    # Initial render
    use_net = current_metric in per_net_metrics and len(all_ssids) > 0
    selected = [] # Começa desmarcado por padrão

    fig, xs, ys, zs, raw_data = _build_heatmap_figure(benchmark_results, flr_path, current_metric, selected, use_net, 
                               show_aps=True)
    canvas_agg = draw_figure(viewer_win["-CANVAS-"].TKCanvas, fig, xs, ys, zs, raw_data, selected)

    while True:
        event, values = viewer_win.read()
        if event in (sg.WIN_CLOSED,):
            break

        needs_redraw = False

        if event in ("-METRIC-", "-UPDATE-"):
            if event == "-METRIC-":
                sel_desc = values["-METRIC-"]
                for i, (key, desc) in enumerate(metric_options):
                    if desc == sel_desc:
                        current_metric = key
                        current_metric_idx = i
                        # Update help tooltip
                        help_text = ConfigurationOptions.configuration.get(key, {}).get("help", "Sem descrição.")
                        viewer_win["-METRIC-HELP-"].set_tooltip(help_text)
                        break
            needs_redraw = True

        if event == "-ALL-":
            for ssid in all_ssids:
                viewer_win[f"-SSID-{ssid}"].update(value=True)

        if event == "-NONE-":
            for ssid in all_ssids:
                viewer_win[f"-SSID-{ssid}"].update(value=False)

        if event.startswith("-HIDE-"):
            ssid_to_hide = event[6:]
            viewer_win[f"-ROW-{ssid_to_hide}"].update(visible=False)
            viewer_win[f"-SSID-{ssid_to_hide}"].update(value=False)

        if event == "-UNHIDE-":
            for ssid in all_ssids:
                viewer_win[f"-ROW-{ssid}"].update(visible=True)

        if event == "-EXPORT-":
            fmt = values["-FMT-"]
            save_path = sg.popup_get_file("Salvar como", save_as=True,
                                          file_types=((fmt.upper(), f"*.{fmt}"),),
                                          default_extension=f".{fmt}")
            if save_path:
                fig.savefig(save_path, format=fmt, dpi=300, bbox_inches='tight')
                sg.popup_ok(f"Imagem salva em:\n{save_path}")

        if event == "-VIEW-SCALE-":
            config = data.get("configuration", {})
            ppm = config.get("pixels_per_meter")
            if not ppm:
                sg.popup_error("Nenhuma calibração de escala encontrada neste projeto.\n"
                               "Calibre a planta no menu 'Novo Projeto' ou 'Continuar Projeto'.",
                               title="Sem Escala")
            else:
                p1 = config.get("scale_p1")
                p2 = config.get("scale_p2")
                meters = config.get("scale_meters")
                
                from matplotlib.pyplot import imread
                from PIL import Image
                try:
                    im = Image.open(flr_path)
                    fdimx, fdimy = im.size
                    
                    fig_scale, ax_scale = plt.subplots(figsize=(10, 8))
                    ax_scale.imshow(imread(flr_path)[::-1], interpolation='bilinear', origin="lower", zorder=1)
                    
                    has_points = (p1 is not None and p2 is not None and meters is not None)
                    if has_points:
                        x1, y1 = p1
                        x2, y2 = p2
                        dist_px = ((x1 - x2)**2 + (y1 - y2)**2)**0.5
                        
                        # Desenha a linha de calibração e os pontos
                        ax_scale.plot([x1, x2], [y1, y2], color='#e74c3c', linestyle='--', linewidth=2.5,
                                      marker='o', markersize=8, markerfacecolor='#f1c40f',
                                      markeredgecolor='black', zorder=10, label="Linha de Calibração")
                        
                        # Rótulos dos pontos
                        ax_scale.text(x1, y1 + fdimy*0.015, "Ponto 1", color='white', fontweight='bold', fontsize=9,
                                      bbox=dict(facecolor='#2c3e50', alpha=0.8, edgecolor='none', boxstyle='round,pad=0.3'), zorder=11)
                        ax_scale.text(x2, y2 + fdimy*0.015, "Ponto 2", color='white', fontweight='bold', fontsize=9,
                                      bbox=dict(facecolor='#2c3e50', alpha=0.8, edgecolor='none', boxstyle='round,pad=0.3'), zorder=11)
                        
                        # Caixa de texto informativa no meio da linha
                        mx = (x1 + x2) / 2
                        my = (y1 + y2) / 2
                        info_text = (
                            f"Distância Real: {meters:.2f} m\n"
                            f"Distância em Pixels: {dist_px:.1f} px\n"
                            f"Escala: {ppm:.2f} px/m"
                        )
                        ax_scale.text(mx, my, info_text, color='white', fontweight='bold', fontsize=10,
                                      ha='center', va='center',
                                      bbox=dict(facecolor='#27ae60', alpha=0.9, edgecolor='white', boxstyle='round,pad=0.5'),
                                      zorder=12)
                        ax_scale.set_title(f"Escala Calibrada: {ppm:.2f} px/m (Pontos Selecionados)", fontsize=12, fontweight='bold', pad=15)
                    else:
                        ax_scale.set_title(f"Escala Calibrada: {ppm:.2f} px/m (Pontos Originais Indisponíveis)", fontsize=12, fontweight='bold', pad=15)
                        sg.popup_ok("Esta planta foi calibrada em uma versão anterior do software.\n"
                                    "Exibindo a barra de escala padrão (5m) de referência.",
                                    title="Aviso")
                    
                    # Desenha barra de escala de 5 metros de referência
                    bar_m = 5.0
                    bar_px = bar_m * ppm
                    
                    margin_x = fdimx * 0.05
                    margin_y = fdimy * 0.05
                    
                    ax_scale.plot([margin_x, margin_x + bar_px], [margin_y, margin_y], color='white', linewidth=4, zorder=20)
                    ax_scale.plot([margin_x, margin_x], [margin_y - fdimy*0.01, margin_y + fdimy*0.01], color='white', linewidth=2, zorder=20)
                    ax_scale.plot([margin_x + bar_px, margin_x + bar_px], [margin_y - fdimy*0.01, margin_y + fdimy*0.01], color='white', linewidth=2, zorder=20)
                    
                    ax_scale.text(margin_x + bar_px/2, margin_y + fdimy*0.015, f"{bar_m:.1f} m", color='white', fontweight='bold',
                                  fontsize=10, ha='center', bbox=dict(facecolor='black', alpha=0.6, edgecolor='none', boxstyle='round,pad=0.2'), zorder=21)
                    
                    ax_scale.axis('off')
                    fig_scale.tight_layout()
                    plt.show()
                    plt.close(fig_scale)
                except Exception as ex:
                    sg.popup_error(f"Erro ao carregar/exibir a escala: {ex}")

        if needs_redraw:
            use_net = current_metric in per_net_metrics and len(all_ssids) > 0
            selected = [s for s in all_ssids if values.get(f"-SSID-{s}", False)]
            plt.close(fig)
            delete_figure(canvas_agg)
            fig, xs, ys, zs, raw_data = _build_heatmap_figure(benchmark_results, flr_path, current_metric, selected, use_net,
                                       show_aps=values["-SHOW-APS-"])
            canvas_agg = draw_figure(viewer_win["-CANVAS-"].TKCanvas, fig, xs, ys, zs, raw_data, selected)

    plt.close(fig)
    viewer_win.close()


def generate_report_dialog():
    import re
    mac_pattern = re.compile(r'\([0-9a-fA-F]{2}(:[0-9a-fA-F]{2}){5}\)$')
    gen_ssids = []
    spec_ssids = []

    layout = [
        [sg.Text("Gerar Relatório de Auditoria (CSV)", font=("Helvetica", 16))],
        [sg.HorizontalSeparator()],
        [sg.Frame("Arquivos", [
            [sg.Text("Configuração (.json):"),
             sg.Input(key="-CFG-", size=(40, 1), enable_events=True),
             sg.FileBrowse("Procurar", file_types=(("JSON", "*.json"),))],
        ], expand_x=True)],
        [sg.Frame("Redes Foco da Análise (Opcional)", [
            [sg.Text("Selecione as redes que deseja detalhar no Inventário de APs:")],
            [
                sg.Column([
                    [sg.Text("Redes por SSID (Geral):", font=("Helvetica", 10, "bold"), text_color="lightblue")],
                    [sg.Listbox(values=[], size=(35, 10), select_mode=sg.LISTBOX_SELECT_MODE_MULTIPLE, key="-SSIDS-GEN-")],
                    [sg.Button("Selecionar Todas", key="-SELECT-ALL-GEN-", size=(15, 1)),
                     sg.Button("Limpar", key="-CLEAR-GEN-", size=(8, 1))]
                ], pad=(5, 5)),
                sg.Column([
                    [sg.Text("Access Points (BSSID Específico):", font=("Helvetica", 10, "bold"), text_color="lightgreen")],
                    [sg.Listbox(values=[], size=(45, 10), select_mode=sg.LISTBOX_SELECT_MODE_MULTIPLE, key="-SSIDS-SPEC-")],
                    [sg.Button("Selecionar Todos", key="-SELECT-ALL-SPEC-", size=(15, 1)),
                     sg.Button("Limpar", key="-CLEAR-SPEC-", size=(8, 1))]
                ], pad=(5, 5))
            ],
            [sg.Text("Dica: Segure CTRL para selecionar várias. Se nenhuma for selecionada em ambas as listas, todas serão incluídas.", font=("Helvetica", 8, "italic"))]
        ], expand_x=True)],
        [sg.Button("Gerar Relatório", key="-GEN-", size=(18, 1)),
         sg.Button("Cancelar", key="-CANCEL-")],
    ]
    window = sg.Window(f"{TITLE} - Relatório", layout, finalize=True)
    while True:
        event, values = window.read()
        if event in (sg.WIN_CLOSED, "-CANCEL-"):
            break
            
        if event == "-CFG-":
            cfg_path = values["-CFG-"].strip()
            if cfg_path and os.path.isfile(cfg_path):
                try:
                    tmp_data = load_json(cfg_path)
                    tmp_results = tmp_data.get("results", {})
                    tmp_ssids = _extract_all_ssids(tmp_results)
                    gen_ssids = sorted([s for s in tmp_ssids if not mac_pattern.search(s)])
                    spec_ssids = sorted([s for s in tmp_ssids if mac_pattern.search(s)])
                    window["-SSIDS-GEN-"].update(values=gen_ssids)
                    window["-SSIDS-SPEC-"].update(values=spec_ssids)
                except Exception as e:
                    pass

        if event == "-SELECT-ALL-GEN-":
            window["-SSIDS-GEN-"].update(set_to_index=list(range(len(gen_ssids))))

        if event == "-CLEAR-GEN-":
            window["-SSIDS-GEN-"].update(set_to_index=[])

        if event == "-SELECT-ALL-SPEC-":
            window["-SSIDS-SPEC-"].update(set_to_index=list(range(len(spec_ssids))))

        if event == "-CLEAR-SPEC-":
            window["-SSIDS-SPEC-"].update(set_to_index=[])
                    
        if event == "-GEN-":
            cfg_path = values["-CFG-"].strip()
            if not cfg_path or not os.path.isfile(cfg_path):
                sg.popup_error("Selecione um .json válido."); continue
            
            data = load_json(cfg_path)
            results = data.get("results", {})
            config = data.get("configuration", {})
            
            selected_gen = values.get("-SSIDS-GEN-", [])
            selected_spec = values.get("-SSIDS-SPEC-", [])
            selected_ssids = list(selected_gen) + list(selected_spec)
            
            # Summary calculation
            all_ssids = _extract_all_ssids(results)
            num_points = len(results)
            
            # Estruturando os dados para Excel (xlsx)
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            import datetime
            
            report_path = cfg_path.replace(".json", "_RELATORIO_EXCEL.xlsx")
            
            # 1. Pré-processa o Inventário de APs para calcular médias agregadas de rádio e métricas por rede (SSID)
            ap_inventory = {}
            bssid_rssis = {}
            bssid_snrs = {}
            ssid_rssis = {}
            ssid_snrs = {}

            for p in results.values():
                nets = p.get("results", {}).get("networks", {})
                for ssid_full, net_data in nets.items():
                    # Usamos apenas BSSIDs específicos para evitar duplicidade de medições nas médias
                    is_bssid = net_data.get("is_bssid_specific", False) or "(" in ssid_full
                    if not is_bssid:
                        continue
                    
                    bssid = net_data.get("ssid_mac", "N/A")
                    if bssid == "N/A":
                        continue
                        
                    clean_ssid = ssid_full.split(" [")[0] if " [" in ssid_full else ssid_full
                    rssi = net_data.get("signal_strength")
                    snr = net_data.get("snr_estimated")
                    
                    if rssi is not None:
                        if bssid not in bssid_rssis:
                            bssid_rssis[bssid] = []
                        bssid_rssis[bssid].append(rssi)
                        
                        if clean_ssid not in ssid_rssis:
                            ssid_rssis[clean_ssid] = []
                        ssid_rssis[clean_ssid].append(rssi)
                        
                    if snr is not None:
                        if bssid not in bssid_snrs:
                            bssid_snrs[bssid] = []
                        bssid_snrs[bssid].append(snr)
                        
                        if clean_ssid not in ssid_snrs:
                            ssid_snrs[clean_ssid] = []
                        ssid_snrs[clean_ssid].append(snr)

                    if bssid not in ap_inventory or net_data["signal_strength"] > ap_inventory[bssid]["signal_strength"]:
                        freq = net_data.get("channel_frequency", 0)
                        band = "5GHz" if freq > 3000 else "2.4GHz"
                        ap_inventory[bssid] = {
                            "ssid": clean_ssid,
                            "channel": net_data.get("channel", 0),
                            "band": band,
                            "signal_strength": net_data.get("signal_strength", 0),
                            "is_secure": "Fechada" if net_data.get("is_secure", False) else "Aberta",
                            "bss_type": net_data.get("bss_type", "N/A"),
                            "vendor": net_data.get("vendor", "N/A"),
                            "station_count": net_data.get("bss_load_station_count", 0),
                            "utilization": net_data.get("bss_load_channel_utilization", 0.0)
                        }

            # Calcula médias
            bssid_avg_rssi = {b: sum(lst)/len(lst) for b, lst in bssid_rssis.items() if lst}
            bssid_avg_snr = {b: sum(lst)/len(lst) for b, lst in bssid_snrs.items() if lst}
            
            # Médias de BSS Load (Apenas APs que reportaram suporte)
            all_utils = [info['utilization'] for info in ap_inventory.values() if info['utilization'] > 0.0 or info['station_count'] > 0]
            all_stas = [info['station_count'] for info in ap_inventory.values() if info['utilization'] > 0.0 or info['station_count'] > 0]
            avg_load_util = sum(all_utils) / len(all_utils) if all_utils else 0.0
            avg_load_sta = sum(all_stas) / len(all_stas) if all_stas else 0.0
            
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Relatório Wi-Fi"
                
                # Helper para escrever linha com formatação
                def append_row(ws, data, bold=False, fill_color=None):
                    ws.append(data)
                    row_idx = ws.max_row
                    for col_idx, cell_value in enumerate(data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if bold:
                            cell.font = Font(bold=True)
                        if fill_color:
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                # Cabeçalho do Projeto
                append_row(ws, ["RELATÓRIO DE AUDITORIA WI-FI"], bold=True)
                ws.append(["Projeto", os.path.basename(cfg_path)])
                ws.append(["Data do Relatório", datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
                ws.append(["Pontos de Medição", num_points])
                ws.append(["Interface", config.get('target_interface', 'N/A')])
                ws.append([])
                
                # Seção 1: Médias
                append_row(ws, ["1. MÉTRICAS MÉDIAS DO SITE SURVEY"], bold=True)
                append_row(ws, ["Métrica", "Valor Médio"], bold=True, fill_color="D9D9D9")
                metrics_to_summary = {
                    "signal_strength": "RSSI Médio (dBm)",
                    "snr_estimated": "SNR Médio (dB)",
                    "ap_density": "Densidade Média de APs",
                    "adjacent_channel_interference": "Interferência Adjacente Média (APs)"
                }
                for key, label in metrics_to_summary.items():
                    vals = [p["results"][key] for p in results.values() if "results" in p and key in p["results"]]
                    if vals:
                        avg = sum(vals) / len(vals)
                        ws.append([label, round(avg, 2)])
                
                # Adiciona as médias de utilização e dispositivos na Seção 1
                if all_utils:
                    ws.append(["Utilização Média do Canal (%)", round(avg_load_util, 2)])
                if all_stas:
                    ws.append(["Média de Clientes Conectados por AP", round(avg_load_sta, 2)])
                
                ws.append([])
                
                # Seção 2: Inventário de APs
                append_row(ws, ["2. INVENTÁRIO DE ACCESS POINTS"], bold=True)
                headers_inv = ["SSID", "BSSID (MAC)", "Canal", "Banda", "Melhor Sinal (dBm)", "RSSI Médio AP (dBm)", "SNR Médio AP (dB)", "Segurança", "Tipo", "Fabricante", "Clientes Conectados", "Utilização do Canal (%)"]
                append_row(ws, headers_inv, bold=True, fill_color="D9D9D9")
                
                for bssid, info in sorted(ap_inventory.items(), key=lambda x: x[1]['ssid']):
                    clean_ssid = info['ssid']
                    candidate_gen = f"{info['ssid']} [{info['band']}]"
                    candidate_spec = f"{info['ssid']} [{info['band']}] ({bssid})"
                    if selected_ssids and (candidate_gen not in selected_ssids and candidate_spec not in selected_ssids):
                        continue
                        
                    avg_rssi_ap = bssid_avg_rssi.get(bssid, 0.0)
                    avg_snr_ap = bssid_avg_snr.get(bssid, 0.0)
                    
                    ws.append([
                        clean_ssid,
                        bssid,
                        info['channel'],
                        info['band'],
                        info['signal_strength'],
                        round(avg_rssi_ap, 1),
                        round(avg_snr_ap, 1),
                        info['is_secure'],
                        info['bss_type'],
                        info['vendor'],
                        info['station_count'],
                        round(info['utilization'], 1)
                    ])
                    
                # Seção 3: Análise de congestionamento por canal
                ws.append([])
                append_row(ws, ["3. ANÁLISE DE CONGESTIONAMENTO POR CANAL"], bold=True)
                headers_cong = ["Canal", "Banda", "APs Detectados", "Utilização Média (%)", "Utilização Máxima (%)", "Média de Clientes Conectados"]
                append_row(ws, headers_cong, bold=True, fill_color="D9D9D9")
                
                channel_stats = {}
                for info in ap_inventory.values():
                    ch = info["channel"]
                    if ch == 0: continue
                    if ch not in channel_stats:
                        channel_stats[ch] = {
                            "band": info["band"],
                            "aps_count": 0,
                            "utils": [],
                            "stas": []
                        }
                    channel_stats[ch]["aps_count"] += 1
                    channel_stats[ch]["utils"].append(info["utilization"])
                    channel_stats[ch]["stas"].append(info["station_count"])
                    
                for ch in sorted(channel_stats.keys()):
                    stat = channel_stats[ch]
                    avg_u = sum(stat["utils"]) / len(stat["utils"]) if stat["utils"] else 0.0
                    max_u = max(stat["utils"]) if stat["utils"] else 0.0
                    avg_s = sum(stat["stas"]) / len(stat["stas"]) if stat["stas"] else 0.0
                    ws.append([
                        ch,
                        stat["band"],
                        stat["aps_count"],
                        round(avg_u, 2),
                        round(max_u, 2),
                        round(avg_s, 2)
                    ])

                # Ajusta tamanho das colunas para melhor visualização
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter # Nome da coluna (A, B, C...)
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width

                wb.save(report_path)

                sg.popup_ok(f"Relatório Excel gerado com sucesso!\n\nSalvo em: {report_path}")
                os.startfile(os.path.dirname(report_path))
            except Exception as e:
                sg.popup_error(f"Erro ao salvar o relatório: {e}")
            break
    window.close()


# ── Main Loop ──────────────────────────────────────────────────────

def main():
    logger.info("Versão: %s", __version__)
    while True:
        choice = main_menu()
        if choice is None:
            logger.info("Aplicação encerrada pelo usuário.")
            break
        elif choice == "-NEW-":
            logger.info("Usuário selecionou: Novo Projeto")
            result = config_wizard()
            if result:
                try:
                    logger.info("Iniciando benchmark: config='%s' map='%s'", result['config'], result['map'])
                    start_gui(result["map"], result["config"])
                except Exception as e:
                    logger.exception("Erro fatal no benchmark: %s", e)
                    sg.popup_error(
                        f"Erro durante o benchmark:\n\n{e}\n\n"
                        f"Verifique o arquivo de log para mais detalhes:\n{LOG_FILE}"
                    )
        elif choice == "-OPEN-":
            logger.info("Usuário selecionou: Continuar Projeto")
            result = open_project()
            if result:
                try:
                    logger.info("Retomando benchmark: config='%s' map='%s'", result['config'], result['map'])
                    start_gui(result["map"], result["config"])
                except Exception as e:
                    logger.exception("Erro fatal no benchmark: %s", e)
                    sg.popup_error(
                        f"Erro durante o benchmark:\n\n{e}\n\n"
                        f"Verifique o arquivo de log para mais detalhes:\n{LOG_FILE}"
                    )
        elif choice == "-PLOT-":
            logger.info("Usuário selecionou: Gerar Mapas de Calor")
            plot_dialog()
        elif choice == "-REPORT-":
            logger.info("Usuário selecionou: Gerar Relatório")
            generate_report_dialog()



if __name__ == "__main__":
    main()
